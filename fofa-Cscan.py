# coding:utf-8
import os
import sys
import json
import base64
import requests
import ipaddress
import tableprint
from config import *
import openpyxl as ws
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


def http(url):
    try:
        r = requests.get(url=url,verify=False)
        return r.text
    except Exception as e:
        raise e

def ip(ip):
    qbase = base64.b64encode(str("ip=%s"%ip).encode('utf-8'))
    res = http('https://fofa.so/api/v1/search/all?email=%s&key=%s&qbase64=%s&size=100&page=1&fields=ip,host,title,port,protocol'%(define.FOFA_EMAIL,define.Apikey,str(qbase,'utf-8')))
    [write_xlsx(i) for i in json.loads(res)["results"]]

def creat_txt():
    if os.path.exists(define.filename) == False:
        if os.path.exists('out/') == False:
            os.mkdir('out')
        with open(define.filename,'a+',encoding='utf-8') as a:
            print(define.BLUE+"[*]创建文件成功 %s"%define.filename)
    else:
        print(define.RED+"[*]文件已存在 文件为:%s"%define.filename)

def write_xlsx(res):
    print(define.GREEN+"[*]内容正在写入 :%s"%res)
    wb = ws.load_workbook(define.filename)
    sheet1 = wb['Sheet']
    num = sheet1.max_row
    sheet1.cell(row = num+1,column = 1,value = res[0])
    sheet1.cell(row = num+1,column = 2,value = res[1])
    sheet1.cell(row = num+1,column = 3,value = res[2])
    sheet1.cell(row = num+1,column = 4,value = res[3])
    sheet1.cell(row = num+1,column = 5,value = res[4])
    wb.save(define.filename)

def creat_xlsx():
    if os.path.exists(define.filename) == False:
        s = 0
        wb = ws.Workbook()
        ws1 = wb.active
        if os.path.exists('out/') == False:
            os.mkdir('out')
        word=['ip','host','title','port','protocol']
        for i in word:
            s = s + 1
            ws1.cell(row =1,column = s,value = i)
        wb.save(define.filename)
        print(define.RED+"[*]创建文件成功 %s"%define.filename)
    else:
        print(define.RED+"[*]文件已存在 文件为:%s"%define.filename)

if __name__ == '__main__':
    print(define.ORANGE+define.banner)

    if len(sys.argv) < 2:
        print(define.ORANGE+define.usage)

    elif sys.argv[1] == '-i':
        print(define.BLUE+"[*]开始抓取")
        creat_xlsx()
        [ip(s) for s in ipaddress.ip_network(str(sys.argv[2]),False)]
        print(define.BLUE+"[*]抓取完毕")

    elif sys.argv[1] == '-f':
        print(define.BLUE+"[*]开始抓取")
        creat_xlsx()
        [ip(s) for x in open(sys.argv[2]) for s in ipaddress.ip_network(x.strip(),False)]
        print(define.BLUE+"[*]抓取完毕")

    else:
        print(define.ORANGE+define.usage)
